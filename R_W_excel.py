# -*- coding: utf-8 -*-
# @Time     : 2020/6/14 20:06
# @Author   : lemon_junjun
# @Email    : 787849486@qq.com
import requests
#用来读取测试用例的
from openpyxl import load_workbook

def read_data(file_name,sheet_name):#读取数据的函数
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    all_case=[]
    for i in range(2,sheet.max_row+1):
        case=[]
        for j in range(1,sheet.max_column-1):
            case.append(sheet.cell(row=i, column=j).value)
        all_case.append(case)
    return all_case#返回所有测试用例数据

if __name__ == '__main__':
    all_case=read_data('test_case.xlsx','recharge')
    print("所有的测试数据为：",all_case)
